import json
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


def test_export_route_returns_excel_file() -> None:
    client = TestClient(app)
    payload = json.dumps(
        {
            "reconciliation_month": "2026-03",
            "platform_name": "ctrip",
            "report_columns": [
                {"key": "product_name", "label": "产品名称", "is_numeric": False},
                {"key": "actual_people", "label": "核销人次", "is_numeric": True},
                {"key": "sales_amount", "label": "销售额", "is_numeric": True},
                {"key": "settlement_paid", "label": "结算实付", "is_numeric": True},
                {"key": "purchase_amount", "label": "采购金额", "is_numeric": True},
                {"key": "profit", "label": "利润", "is_numeric": True},
            ],
            "rows": [
                {
                    "product_name": "测试产品",
                    "metrics": {
                        "actual_people": 3,
                        "sales_amount": 150.0,
                        "settlement_paid": 150.0,
                        "purchase_amount": 120.0,
                        "profit": 30.0,
                    },
                }
            ],
        },
        ensure_ascii=False,
    )

    response = client.post("/export", data={"payload": payload})

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["对账结果"]

    assert worksheet["A1"].value == "对账月份"
    assert worksheet["B1"].value == "2026-03"
    assert worksheet["A2"].value == "平台"
    assert worksheet["B2"].value == "携程"
    assert worksheet["A4"].value == "产品名称"
    assert worksheet["B4"].value == "核销人次"
    assert worksheet["C4"].value == "销售额"
    assert worksheet["A5"].value == "测试产品"
    assert worksheet["B5"].value == 3
    assert worksheet["C5"].value == 150
    assert worksheet["F5"].value == 30
    assert workbook.sheetnames == ["对账结果"]


def test_export_route_supports_meituan_dynamic_columns() -> None:
    client = TestClient(app)
    payload = json.dumps(
        {
            "reconciliation_month": "2026-03",
            "platform_name": "meituan",
            "report_columns": [
                {"key": "product_name", "label": "产品名称", "is_numeric": False},
                {"key": "actual_people", "label": "核销人次", "is_numeric": True},
                {"key": "sales_amount", "label": "销售额", "is_numeric": True},
                {"key": "technical_service_fee", "label": "技术服务费", "is_numeric": True},
                {"key": "merchant_coupon", "label": "优惠券（商家承担）", "is_numeric": True},
                {"key": "settlement_paid", "label": "结算实付", "is_numeric": True},
                {"key": "purchase_amount", "label": "采购金额", "is_numeric": True},
                {"key": "profit", "label": "利润", "is_numeric": True},
            ],
            "rows": [
                {
                    "product_name": "测试产品",
                    "metrics": {
                        "actual_people": 3,
                        "sales_amount": 150.0,
                        "technical_service_fee": 12.0,
                        "merchant_coupon": 5.0,
                        "settlement_paid": 140.0,
                        "purchase_amount": 120.0,
                        "profit": 20.0,
                    },
                }
            ],
        },
        ensure_ascii=False,
    )

    response = client.post("/export", data={"payload": payload})

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["对账结果"]

    assert worksheet["B2"].value == "美团"
    assert worksheet["D4"].value == "技术服务费"
    assert worksheet["E4"].value == "优惠券（商家承担）"
    assert worksheet["H4"].value == "利润"
    assert worksheet["D5"].value == 12
    assert worksheet["E5"].value == 5
    assert worksheet["H5"].value == 20


def test_export_differences_route_returns_excel_file() -> None:
    client = TestClient(app)
    payload = json.dumps(
        {
            "reconciliation_month": "2026-03",
            "platform_name": "ctrip",
            "internal_only_order_nos": ["JTX-001", "JTX-002"],
            "external_only_order_nos": ["CTRIP-404"],
        },
        ensure_ascii=False,
    )

    response = client.post("/export-differences", data={"payload": payload})

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["聚天下有、第三方当月无", "第三方有、聚天下当月无"]

    internal_only_sheet = workbook["聚天下有、第三方当月无"]
    assert internal_only_sheet["A1"].value == "对账月份"
    assert internal_only_sheet["B1"].value == "2026-03"
    assert internal_only_sheet["A2"].value == "平台"
    assert internal_only_sheet["B2"].value == "携程"
    assert internal_only_sheet["A4"].value == "序号"
    assert internal_only_sheet["B4"].value == "订单号"
    assert internal_only_sheet["A5"].value == 1
    assert internal_only_sheet["B5"].value == "JTX-001"
    assert internal_only_sheet["A6"].value == 2
    assert internal_only_sheet["B6"].value == "JTX-002"

    external_only_sheet = workbook["第三方有、聚天下当月无"]
    assert external_only_sheet["A1"].value == "对账月份"
    assert external_only_sheet["B1"].value == "2026-03"
    assert external_only_sheet["A2"].value == "平台"
    assert external_only_sheet["B2"].value == "携程"
    assert external_only_sheet["A4"].value == "序号"
    assert external_only_sheet["B4"].value == "第三方单号"
    assert external_only_sheet["A5"].value == 1
    assert external_only_sheet["B5"].value == "CTRIP-404"
