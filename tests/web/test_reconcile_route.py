from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


def _find_workbook_by_sheet(sheet_name: str) -> Path:
    base = Path(__file__).resolve().parents[2] / "test_data"
    for path in base.glob("*.xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet_name in workbook.sheetnames:
            return path
    raise FileNotFoundError(f"Could not find workbook containing sheet: {sheet_name}")


def test_reconcile_route_renders_summary_and_table() -> None:
    client = TestClient(app)
    jutianxia_file = _find_workbook_by_sheet("订单列表")
    ctrip_file = _find_workbook_by_sheet("流水")

    with jutianxia_file.open("rb") as jutianxia_stream, ctrip_file.open("rb") as ctrip_stream:
        response = client.post(
            "/reconcile",
            data={
                "reconciliation_month": "2026-03",
                "platform": "ctrip",
            },
            files={
                "jutianxia_file": (
                    jutianxia_file.name,
                    jutianxia_stream,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                "platform_file": (
                    ctrip_file.name,
                    ctrip_stream,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

    assert response.status_code == 200
    assert "成功匹配订单数" in response.text
    assert "1908" in response.text
    assert "汇总产品数" in response.text
    assert "29" in response.text
    assert "被过滤的非当月平台记录数" in response.text
    assert ">3<" in response.text
    assert "结果总览" in response.text
    assert "主结果表" in response.text
    assert "订单差异检查" in response.text
    assert "聚天下有、第三方当月无" in response.text
    assert ">172<" in response.text
    assert "第三方有、聚天下当月无" in response.text
    assert response.text.count(">序号<") == 2
    assert response.text.count(">订单号<") >= 1
    assert response.text.count(">第三方单号<") >= 1
    assert ">1<" in response.text
    assert "14347370" in response.text
    assert response.text.count("<details") == 2
    assert "【即买即用】【线下扫码】东运旅行卢宅景区成人票+卢宅文创产品1份" in response.text
    assert "导出 Excel" in response.text
