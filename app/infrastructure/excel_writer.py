"""Excel 导出基础设施。

该模块把对账结果行写入标准工作簿，并返回二进制字节流供 Web 层下载。
"""

from io import BytesIO

from openpyxl import Workbook


def build_reconciliation_workbook(
    reconciliation_month: str,
    platform_label: str,
    report_columns: list[dict],
    rows: list[dict],
) -> bytes:
    """构建“对账结果”工作簿并返回 .xlsx 字节。

    rows 来自 Web 层序列化后的结果行，字段键需与页面展示保持一致。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "对账结果"

    # 文件头部放基本元信息，便于财务导出后追溯来源。
    worksheet["A1"] = "对账月份"
    worksheet["B1"] = reconciliation_month
    worksheet["A2"] = "平台"
    worksheet["B2"] = platform_label

    # 第 4 行为表头，前 1-3 行预留给元信息。
    for column_index, column in enumerate(report_columns, start=1):
        worksheet.cell(row=4, column=column_index, value=column["label"])

    # 从第 5 行开始逐行写入汇总结果。
    for row_index, row in enumerate(rows, start=5):
        for column_index, column in enumerate(report_columns, start=1):
            value = (
                row["product_name"]
                if column["key"] == "product_name"
                else row["metrics"].get(column["key"], 0)
            )
            worksheet.cell(row=row_index, column=column_index, value=value)

    # 使用内存缓冲区避免落地临时文件，适合 HTTP 直接返回。
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_difference_workbook(
    reconciliation_month: str,
    platform_label: str,
    internal_difference_label: str,
    external_difference_label: str,
    internal_only_order_nos: list[str],
    external_only_order_nos: list[str],
) -> bytes:
    """构建“订单差异”工作簿并返回 .xlsx 字节。"""
    workbook = Workbook()
    internal_only_sheet = workbook.active
    internal_only_sheet.title = "聚天下有、第三方当月无"
    _write_difference_sheet(
        worksheet=internal_only_sheet,
        reconciliation_month=reconciliation_month,
        platform_label=platform_label,
        number_header=internal_difference_label,
        order_nos=internal_only_order_nos,
    )

    external_only_sheet = workbook.create_sheet("第三方有、聚天下当月无")
    _write_difference_sheet(
        worksheet=external_only_sheet,
        reconciliation_month=reconciliation_month,
        platform_label=platform_label,
        number_header=external_difference_label,
        order_nos=external_only_order_nos,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_difference_sheet(
    worksheet,
    reconciliation_month: str,
    platform_label: str,
    number_header: str,
    order_nos: list[str],
) -> None:
    """把一个方向的订单差异写入单个工作表。"""
    worksheet["A1"] = "对账月份"
    worksheet["B1"] = reconciliation_month
    worksheet["A2"] = "平台"
    worksheet["B2"] = platform_label
    worksheet["A4"] = "序号"
    worksheet["B4"] = number_header

    if order_nos:
        for row_index, order_no in enumerate(order_nos, start=5):
            worksheet.cell(row=row_index, column=1, value=row_index - 4)
            worksheet.cell(row=row_index, column=2, value=order_no)
    else:
        worksheet["A5"] = 1
        worksheet["B5"] = "无"
